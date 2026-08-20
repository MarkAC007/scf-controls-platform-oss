"""First-ever tests for scripts/extract_scf_data.py (WP1a, plan §4.7).

Synthetic openpyxl workbooks cover both sheet-name eras the extractor handles:

- the 2025.x era — 'Authoritative Sources' names sheet with 'Mapping Column
  Header' / 'Authoritative Source - Law, Regulation or Framework (LRF)'
  columns;
- the 2026.2 era — the sheet renamed to 'Focal Documents' with 'SCF Column
  Header' / 'Focal Document Name (FDN)' columns.

The framework-name regression tests pin the WP1a fix: display names must come
from the era's name column, never silently degrade to the column-header
fallback (the pre-fix behavior on 2026.2 workbooks, plan §2 known defect).

Fixture identifiers deliberately use a letter after the hyphen (``GOV-A1``,
``E-GOV-A1``) rather than real SCF numbering: the extractor treats them as
opaque strings, and literal control-ID-shaped tokens cannot be written to this
repo (ContainmentGuard).
"""
from __future__ import annotations

import json
import sys
from pathlib import Path

import pytest
from openpyxl import Workbook

# Repo-root scripts/ in a source checkout; /app/scripts in the backend
# container (the backend dir mounts at /app, so parents[2] resolves to /).
_SCRIPTS_CANDIDATES = [
    Path(__file__).resolve().parents[2] / "scripts",
    Path("/app/scripts"),
]
SCRIPTS_DIR = next((p for p in _SCRIPTS_CANDIDATES if p.is_dir()), None)
if SCRIPTS_DIR is None:
    raise RuntimeError(
        f"extract_scf_data.py location not found; tried {_SCRIPTS_CANDIDATES}"
    )
if str(SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_DIR))

import extract_scf_data as extractor  # noqa: E402


# ---------------------------------------------------------------------------
# Synthetic workbook builder (shared with test_catalog_diff.py)
# ---------------------------------------------------------------------------

AICPA_HEADER = "AICPA\nTSC 2017:2022 (used for SOC 2)"
GDPR_HEADER = "GDPR\nEU General Data Protection Regulation"
AICPA_SLUG = extractor.normalize_framework_id(AICPA_HEADER)
GDPR_SLUG = extractor.normalize_framework_id(GDPR_HEADER)

AICPA_NAME_OLD = "AICPA Trust Services Criteria (TSC) 2017"
AICPA_NAME_NEW = (
    "American Institute of Certified Public Accountants (AICPA) "
    "Trust Services Criteria (2017)"
)

# Era-renamed column headers (2026.2 renamed the CMM and domains-principle
# columns alongside the framework-names sheet).
CMM0_HEADER_OLD = "C|P-CMM 0\nNot Performed"
CMM0_HEADER_NEW = "SCR-CMM Level 0\nNot Performed"
PRINCIPLE_HEADER_OLD = "Cybersecurity & Data Privacy by Design (C|P) Principles"
PRINCIPLE_HEADER_NEW = "Security, Compliance & Resilience (SCR) Principles"


def controls_headers(era: str) -> list:
    return [
        "SCF Domain",
        "SCF Control",
        "SCF #",
        "Secure Controls Framework (SCF)\nControl Description",
        "Conformity Validation Cadence",
        "Evidence Request List (ERL) #",
        "SCF Control Question",
        "Relative Control Weighting",
        "PPTDF\nApplicability",
        "NIST CSF\nFunction Grouping",
        CMM0_HEADER_OLD if era == "authoritative_sources" else CMM0_HEADER_NEW,
        AICPA_HEADER,
        GDPR_HEADER,
    ]

DEFAULT_CONTROL_ROWS = [
    (
        "Cybersecurity & Data Protection Governance",
        "Cybersecurity & Data Protection Governance Program",
        "GOV-A1",
        "Mechanisms exist to facilitate the implementation of a governance program.",
        "Annual",
        "E-GOV-A1, E-GOV-A2",
        "Does the organization facilitate a governance program?",
        10,
        "Process",
        "Govern",
        "Practices are non-existent.",
        "CC1.1\nCC1.2",
        "Art 32",
    ),
    (
        "Cybersecurity & Data Protection Governance",
        "Publishing Cybersecurity & Data Protection Documentation",
        "GOV-A2",
        "Mechanisms exist to establish and publish documentation.",
        "Annual",
        "E-GOV-A3",
        None,
        8,
        "Process",
        "Govern",
        None,
        "CC5.3",
        None,
    ),
]

def domains_headers(era: str) -> list:
    return [
        "#",
        "SCF Domain",
        "SCF Identifier",
        PRINCIPLE_HEADER_OLD if era == "authoritative_sources" else PRINCIPLE_HEADER_NEW,
        "Principle Intent",
    ]

DEFAULT_DOMAIN_ROWS = [
    (
        1,
        "Cybersecurity & Data Protection Governance",
        "GOV",
        "Execute a documented, risk-based program.",
        "Organizations specify the development of a program.",
    ),
]

EVIDENCE_HEADERS = [
    "#",
    "ERL #",
    "Area of Focus",
    "Documentation Artifact",
    "Artifact Description",
    "SCF Control Mappings",
]

DEFAULT_EVIDENCE_ROWS = [
    (
        1,
        "E-GOV-A1",
        "Governance",
        "Cybersecurity Program Charter",
        "Charter for the cybersecurity program.",
        "GOV-A1, GOV-A2",
    ),
    (
        2,
        "E-GOV-A2",
        "Governance",
        "Steering Committee Minutes",
        "Minutes evidencing oversight.",
        "GOV-A1",
    ),
]

AO_HEADERS = [
    extractor.COL_AO_SCF_ID,
    extractor.COL_AO_ID,
    extractor.COL_AO_TEXT,
    "PPTDF\nApplicability",
    extractor.COL_AO_ORIGINS,
    "Assessment\nRigor (AR)",
    extractor.COL_AO_PROCEDURE,
    extractor.COL_AO_EXPECTED,
]

DEFAULT_AO_ROWS = [
    (
        "GOV-A1",
        "GOV-A1.1",
        "the organization facilitates a governance program.",
        "Process",
        "SCF",
        3,
        "Examine the program charter.",
        "A charter exists and is approved.",
    ),
]

NAMES_SHEET_OLD = (
    "Authoritative Sources",
    [
        "Geography",
        "Mapping Column Header",
        "Source",
        "Authoritative Source - Law, Regulation or Framework (LRF)",
    ],
    [("General", AICPA_HEADER, "AICPA", AICPA_NAME_OLD)],
)

NAMES_SHEET_NEW = (
    "Focal Documents",
    [
        "Geography",
        "SCF Column Header",
        "Focal Document Identifier (FDI)",
        "Source",
        "Focal Document Name (FDN)",
    ],
    [("General", AICPA_HEADER, "general-aicpa-tsc-2017", "AICPA", AICPA_NAME_NEW)],
)


def _add_sheet(wb: Workbook, title: str, headers: list, rows: list) -> None:
    ws = wb.create_sheet(title)
    ws.append(headers)
    for row in rows:
        ws.append(row)


def build_workbook(
    path,
    version: str = "2026.2",
    era: str = "focal_documents",
    *,
    controls_sheet_title: str | None = None,
    control_rows: list | None = None,
    names_sheet: tuple | None = None,
) -> Path:
    """Write a small synthetic SCF workbook covering all extractor sheets.

    ``era`` selects the framework-names sheet layout: ``authoritative_sources``
    (2025.x) or ``focal_documents`` (2026.2).
    """
    wb = Workbook()
    wb.remove(wb.active)

    _add_sheet(wb, "SCF Domains & Principles", domains_headers(era), DEFAULT_DOMAIN_ROWS)

    if names_sheet is None:
        names_sheet = (
            NAMES_SHEET_OLD if era == "authoritative_sources" else NAMES_SHEET_NEW
        )
    _add_sheet(wb, *names_sheet)

    _add_sheet(
        wb,
        controls_sheet_title or f"SCF {version}",
        controls_headers(era),
        control_rows if control_rows is not None else DEFAULT_CONTROL_ROWS,
    )
    _add_sheet(
        wb, f"Evidence Request List {version}", EVIDENCE_HEADERS, DEFAULT_EVIDENCE_ROWS
    )
    _add_sheet(wb, f"Assessment Objectives {version}", AO_HEADERS, DEFAULT_AO_ROWS)

    wb.save(path)
    return Path(path)


def _extract(tmp_path, **build_kwargs):
    tmp_path.mkdir(parents=True, exist_ok=True)
    workbook = build_workbook(tmp_path / "scf.xlsx", **build_kwargs)
    out_dir = tmp_path / "out"
    meta = extractor.extract_to_dir(workbook, out_dir)
    return meta, out_dir


def _load(out_dir, filename):
    with open(Path(out_dir) / filename) as f:
        return json.load(f)


# ---------------------------------------------------------------------------
# Full extraction, both eras
# ---------------------------------------------------------------------------


def test_extracts_old_era_workbook(tmp_path):
    meta, out_dir = _extract(tmp_path, version="2025.4", era="authoritative_sources")

    assert meta["catalog_version"] == "2025.4"
    assert meta["controls"] == 2
    assert meta["domains"] == 1
    assert meta["evidence"] == 2
    assert meta["assessment_objectives"] == 1

    controls = {c["scf_id"]: c for c in _load(out_dir, "control_guidance.json")["controls"]}
    first = controls["GOV-A1"]
    assert first["control_name"] == "Cybersecurity & Data Protection Governance Program"
    assert first["control_weighting"] == 10
    assert first["pptdf_applicability"] == {
        "people": False,
        "process": True,
        "technology": False,
        "data": False,
        "facility": False,
    }
    assert first["evidence_requests"] == ["E-GOV-A1", "E-GOV-A2"]
    assert first["framework_mappings"][AICPA_SLUG] == ["CC1.1", "CC1.2"]
    assert first["framework_mappings"][GDPR_SLUG] == ["Art 32"]
    assert first["cmm_maturity"] == {"level_0": "Practices are non-existent."}
    assert controls["GOV-A2"]["control_question"] is None
    assert controls["GOV-A2"]["cmm_maturity"] is None

    domains = _load(out_dir, "domains.json")
    assert domains == [
        {
            "order": 1,
            "name": "Cybersecurity & Data Protection Governance",
            "identifier": "GOV",
            "principle": "Execute a documented, risk-based program.",
            "principle_intent": "Organizations specify the development of a program.",
        }
    ]

    erl = _load(out_dir, "erl.json")
    assert set(erl) == {"E-GOV-A1", "E-GOV-A2"}
    assert erl["E-GOV-A1"]["control_mappings"] == ["GOV-A1", "GOV-A2"]

    objectives = _load(out_dir, "assessment_objectives.json")["objectives"]
    assert objectives[0]["ao_id"] == "GOV-A1.1"
    assert objectives[0]["scf_id"] == "GOV-A1"
    assert objectives[0]["assessment_rigor"] == 3
    assert objectives[0]["pptdf_applicability"]["process"] is True


def test_extracts_new_era_workbook(tmp_path):
    meta, out_dir = _extract(tmp_path, version="2026.2", era="focal_documents")

    assert meta["catalog_version"] == "2026.2"
    assert meta["controls"] == 2
    controls = {c["scf_id"]: c for c in _load(out_dir, "control_guidance.json")["controls"]}
    assert controls["GOV-A1"]["framework_mappings"][AICPA_SLUG] == ["CC1.1", "CC1.2"]


def test_cmm_extracted_from_renamed_2026_2_columns(tmp_path):
    """Pre-fix, 2026.2 workbooks silently yielded cmm_maturity=None for every
    control: the extractor only knew the 'C|P-CMM N ...' headers, which were
    renamed to 'SCR-CMM Level N ...' in 2026.2. Both eras must extract."""
    for version, era in (("2025.4", "authoritative_sources"), ("2026.2", "focal_documents")):
        _, out_dir = _extract(tmp_path / version.replace(".", "_"), version=version, era=era)
        controls = {
            c["scf_id"]: c for c in _load(out_dir, "control_guidance.json")["controls"]
        }
        assert controls["GOV-A1"]["cmm_maturity"] == {
            "level_0": "Practices are non-existent."
        }, f"cmm_maturity lost on {era}"


def test_domain_principle_extracted_from_renamed_2026_2_column(tmp_path):
    """Pre-fix, 2026.2 workbooks silently yielded principle='' for every
    domain: the principles column was renamed 'Cybersecurity & Data Privacy by
    Design (C|P) Principles' -> 'Security, Compliance & Resilience (SCR)
    Principles'. Both eras must extract."""
    for version, era in (("2025.4", "authoritative_sources"), ("2026.2", "focal_documents")):
        _, out_dir = _extract(tmp_path / version.replace(".", "_"), version=version, era=era)
        domains = _load(out_dir, "domains.json")
        assert domains[0]["principle"] == "Execute a documented, risk-based program.", (
            f"principle lost on {era}"
        )
        assert domains[0]["principle_intent"] == (
            "Organizations specify the development of a program."
        )


def test_cmm_candidate_headers_pin_both_eras():
    """Pin the generated candidate lists against the literal header names
    verified in the real 2025.4 and 2026.2 workbooks (all six levels)."""
    assert [c[0] for c in extractor.COL_CMM_CANDIDATES] == [
        "C|P-CMM 0 Not Performed",
        "C|P-CMM 1 Performed Informally",
        "C|P-CMM 2 Planned & Tracked",
        "C|P-CMM 3 Well Defined",
        "C|P-CMM 4 Quantitatively Controlled",
        "C|P-CMM 5 Continuously Improving",
    ]
    assert [c[1] for c in extractor.COL_CMM_CANDIDATES] == [
        "SCR-CMM Level 0 Not Performed",
        "SCR-CMM Level 1 Performed Informally",
        "SCR-CMM Level 2 Planned & Tracked",
        "SCR-CMM Level 3 Well Defined",
        "SCR-CMM Level 4 Quantitatively Controlled",
        "SCR-CMM Level 5 Continuously Improving",
    ]


def test_domains_missing_principle_column_warns_and_extracts_empty(tmp_path, capsys):
    """A future rename degrades to principle='' with a warning, not a crash."""
    import pandas as pd

    sheet = "SCF Domains & Principles"
    wb = Workbook()
    wb.remove(wb.active)
    mangled_headers = [
        "#",
        "SCF Domain",
        "SCF Identifier",
        "Some Future Principles Column",
        "Principle Intent",
    ]
    _add_sheet(wb, sheet, mangled_headers, DEFAULT_DOMAIN_ROWS)
    path = tmp_path / "mangled.xlsx"
    wb.save(path)

    domains = extractor.extract_domains(pd.ExcelFile(path), sheet)
    assert domains[0]["principle"] == ""
    assert domains[0]["name"] == "Cybersecurity & Data Protection Governance"
    assert "no recognised principles column" in capsys.readouterr().out


def test_resolve_catalog_sheets_both_eras(tmp_path):
    import pandas as pd

    old = build_workbook(tmp_path / "old.xlsx", version="2025.4", era="authoritative_sources")
    new = build_workbook(tmp_path / "new.xlsx", version="2026.2", era="focal_documents")

    old_sheets = extractor.resolve_catalog_sheets(pd.ExcelFile(old))
    assert old_sheets["catalog_version"] == "2025.4"
    assert old_sheets["authoritative_sources"] == "Authoritative Sources"

    new_sheets = extractor.resolve_catalog_sheets(pd.ExcelFile(new))
    assert new_sheets["catalog_version"] == "2026.2"
    assert new_sheets["authoritative_sources"] == "Focal Documents"
    assert new_sheets["evidence"] == "Evidence Request List 2026.2"
    assert new_sheets["assessment_objectives"] == "Assessment Objectives 2026.2"


# ---------------------------------------------------------------------------
# Framework display names — regression tests pinning the WP1a fix
# ---------------------------------------------------------------------------


def test_framework_names_old_era_from_lrf_column(tmp_path):
    _, out_dir = _extract(tmp_path, version="2025.4", era="authoritative_sources")
    names = _load(out_dir, "frameworks.json")
    assert names[AICPA_SLUG] == AICPA_NAME_OLD


def test_framework_names_new_era_from_fdn_column(tmp_path):
    """Pre-fix, 2026.2 workbooks silently degraded every framework display
    name to the column-header fallback because the extractor read the old
    era's column names. The FDN column must win."""
    _, out_dir = _extract(tmp_path, version="2026.2", era="focal_documents")
    names = _load(out_dir, "frameworks.json")
    assert names[AICPA_SLUG] == AICPA_NAME_NEW
    # And explicitly: NOT the cleaned column-header fallback.
    assert names[AICPA_SLUG] != extractor.clean_column_name(AICPA_HEADER)


def test_framework_names_fallback_for_unlisted_column(tmp_path):
    """A framework column with no row in the names sheet keeps the cleaned
    column header as its display name (existing fallback, both eras)."""
    for version, era in (("2025.4", "authoritative_sources"), ("2026.2", "focal_documents")):
        _, out_dir = _extract(tmp_path / version.replace(".", "_"), version=version, era=era)
        names = _load(out_dir, "frameworks.json")
        assert names[GDPR_SLUG] == extractor.clean_column_name(GDPR_HEADER)


def test_framework_names_unrecognised_sheet_falls_back(tmp_path, capsys):
    """A names sheet with neither era's columns degrades to header fallbacks —
    but now announces it instead of failing silently."""
    names_sheet = (
        "Focal Documents",
        ["Geography", "Some Future Column", "Another Column"],
        [("General", "AICPA", "whatever")],
    )
    _, out_dir = _extract(
        tmp_path, version="2026.2", era="focal_documents", names_sheet=names_sheet
    )
    names = _load(out_dir, "frameworks.json")
    assert names[AICPA_SLUG] == extractor.clean_column_name(AICPA_HEADER)
    assert names[GDPR_SLUG] == extractor.clean_column_name(GDPR_HEADER)
    assert "no recognised framework-name columns" in capsys.readouterr().out


def test_framework_name_blank_falls_back_to_header(tmp_path):
    """A names row whose name cell is blank falls back to its column header."""
    names_sheet = (
        "Focal Documents",
        NAMES_SHEET_NEW[1],
        [("General", AICPA_HEADER, "general-aicpa-tsc-2017", "AICPA", None)],
    )
    _, out_dir = _extract(
        tmp_path, version="2026.2", era="focal_documents", names_sheet=names_sheet
    )
    names = _load(out_dir, "frameworks.json")
    assert names[AICPA_SLUG] == extractor.clean_column_name(AICPA_HEADER)


# ---------------------------------------------------------------------------
# Version detection
# ---------------------------------------------------------------------------


def test_unrecognisable_workbook_raises(tmp_path):
    workbook = build_workbook(
        tmp_path / "bad.xlsx", version="2026.2", controls_sheet_title="SCF Controls"
    )
    with pytest.raises(ValueError, match="catalog version"):
        extractor.extract_to_dir(workbook, tmp_path / "out")
